from __future__ import annotations

import hashlib
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import Workbook

from versevad.adapters.concreteness import (
    BrysbaertConcretenessAdapter,
    ConcretenessAdapterError,
)
from versevad.core import AnalysisModule, ModuleInput, ResourceSpec, ResourceState
from versevad.lexical_semantic.concreteness import (
    BRYSBAERT_CONCRETENESS_FILENAME,
    BRYSBAERT_CONCRETENESS_SHA256,
    ConcretenessConfiguration,
    ConcretenessMatchMethod,
    ConcretenessModule,
)
from versevad.preprocessing import create_text_document


HEADER = (
    "Word",
    "Bigram",
    "Conc.M",
    "Conc.SD",
    "Unknown",
    "Total",
    "Percent_known",
    "SUBTLEX",
)


def _write_workbook(path: Path, rows: list[tuple[object, ...]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(HEADER)
    for row in rows:
        sheet.append(row)
    workbook.create_sheet("Sheet2")
    workbook.create_sheet("Sheet3")
    workbook.save(path)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _module(
    tmp_path: Path,
    rows: list[tuple[object, ...]],
) -> ConcretenessModule:
    source = tmp_path / "ratings.xlsx"
    _write_workbook(source, rows)
    spec = ResourceSpec(
        resource_id="synthetic-concreteness",
        display_name="Synthetic concreteness fixture",
        relative_path=source.name,
        version="synthetic-v1",
        accepted_sha256=(_sha256(source),),
        citation="Constructed test fixture.",
        license_notice="Synthetic test data.",
    )
    return ConcretenessModule(tmp_path, resource_spec=spec)


def _analyze(
    module: ConcretenessModule,
    preprocessor,
    text: str,
    *,
    configuration: ConcretenessConfiguration | None = None,
):
    poem = preprocessor.process_document(
        create_text_document("concreteness-test", "Concreteness test", text)
    )
    return module.analyze_detailed(
        ModuleInput.from_poem_document(poem),
        configuration or ConcretenessConfiguration(),
    )


def test_adapter_preserves_source_and_validates_expected_columns(tmp_path: Path) -> None:
    source = tmp_path / "ratings.xlsx"
    _write_workbook(
        source,
        [
            ("stone", 0, 4.9, 0.2, 0, 30, 1.0, 100),
            ("dark night", 1, 3.8, 0.8, 1, 30, 29 / 30, 2),
        ],
    )
    before = _sha256(source)

    lexicon = BrysbaertConcretenessAdapter().load(source)

    assert _sha256(source) == before
    assert lexicon.validation.is_valid
    assert lexicon.validation.total_rows == 2
    assert lexicon.validation.usable_entries == 2
    assert lexicon.validation.phrase_entries == 1
    assert lexicon.validation.source_sha256 == before
    assert lexicon.entries["stone"].mean == pytest.approx(4.9)
    assert lexicon.entries["dark night"].is_multiword
    assert lexicon.entries["dark night"].rater_count == 30


def test_adapter_rejects_missing_columns_duplicates_and_bad_ranges(
    tmp_path: Path,
) -> None:
    missing = tmp_path / "missing.xlsx"
    workbook = Workbook()
    workbook.active.title = "Sheet1"
    workbook.active.append(("Word", "Conc.M"))
    workbook.active.append(("stone", 4.0))
    workbook.save(missing)
    with pytest.raises(ConcretenessAdapterError, match="expected columns"):
        BrysbaertConcretenessAdapter().load(missing)

    malformed = tmp_path / "malformed.xlsx"
    _write_workbook(
        malformed,
        [
            ("stone", 0, 4.0, 0.2, 0, 30, 1.0, 100),
            ("STONE", 0, 6.0, -1.0, 31, 30, 1.2, -1),
        ],
    )
    with pytest.raises(ConcretenessAdapterError) as captured:
        BrysbaertConcretenessAdapter().load(malformed)
    detail = captured.value.technical_detail.casefold()
    assert "duplicate" in detail
    assert "outside" in detail
    assert not captured.value.data_changed


def test_exact_precedes_lemma_and_unmatched_values_stay_missing(
    tmp_path: Path,
    preprocessor,
) -> None:
    module = _module(
        tmp_path,
        [
            ("stone", 0, 5.0, 0.0, 0, 30, 1.0, 100),
            ("idea", 0, 1.0, 0.0, 0, 30, 1.0, 100),
            ("saw", 0, 2.0, 0.0, 0, 30, 1.0, 100),
            ("see", 0, 4.0, 0.0, 0, 30, 1.0, 100),
        ],
    )
    result = _analyze(
        module,
        preprocessor,
        "stone stones\nidea\nidea\nquorvax",
        configuration=ConcretenessConfiguration(
            low_coverage_warning_threshold=0.9
        ),
    )
    by_surface = {}
    for row in result.token_audit:
        by_surface.setdefault(row.surface_form, []).append(row)

    assert by_surface["stone"][0].match_method is ConcretenessMatchMethod.EXACT
    assert by_surface["stones"][0].match_method is ConcretenessMatchMethod.LEMMA
    assert by_surface["quorvax"][0].match_method is ConcretenessMatchMethod.UNMATCHED
    assert by_surface["quorvax"][0].rating is None
    assert result.summary.eligible_token_count == 5
    assert result.summary.rated_token_count == 4
    assert result.summary.token_coverage == pytest.approx(0.8)
    assert result.summary.eligible_unique_type_count == 4
    assert result.summary.rated_unique_type_count == 3
    assert result.summary.unique_type_coverage == pytest.approx(0.75)
    assert result.summary.statistics.mean == pytest.approx(3.0)
    assert result.summary.statistics.median == pytest.approx(3.0)
    assert result.summary.statistics.population_standard_deviation == pytest.approx(2.0)
    assert result.summary.interquartile_range == pytest.approx(4.0)
    assert result.summary.highly_concrete_proportion == pytest.approx(0.5)
    assert result.summary.highly_abstract_proportion == pytest.approx(0.5)
    idea = next(term for term in result.term_summaries if term.source_term == "idea")
    assert idea.rated_token_occurrences == 2
    assert any(warning.code == "low_coverage" for warning in result.module_result.warnings)


def test_alphabetic_number_word_matches_but_numeric_literal_is_ineligible(
    tmp_path: Path,
    preprocessor,
) -> None:
    result = _analyze(
        _module(tmp_path, [("one", 0, 2.5, 0.5, 0, 30, 1.0, 100)]),
        preprocessor,
        "one 27",
    )
    by_surface = {row.surface_form: row for row in result.token_audit}

    assert by_surface["one"].included is True
    assert by_surface["one"].rating == pytest.approx(2.5)
    assert "alphabetically spelled" in by_surface["one"].reason
    assert by_surface["27"].eligible is False
    assert "pure numeric literal" in by_surface["27"].reason


def test_exact_source_form_is_not_replaced_by_a_different_lemma_entry(
    tmp_path: Path,
    preprocessor,
) -> None:
    module = _module(
        tmp_path,
        [
            ("saw", 0, 2.0, 0.5, 0, 30, 1.0, 10),
            ("see", 0, 4.0, 0.5, 0, 30, 1.0, 10),
        ],
    )
    poem = preprocessor.process_document(
        create_text_document("exact-first", "Exact first", "saw")
    )
    token = replace(
        poem.tokens[0],
        lemma="see",
        normalized_lemma="see",
        part_of_speech="VERB",
    )
    module_input = ModuleInput(
        document=poem.source,
        tokens=(token,),
        preprocessing=poem.preprocessing,
    )

    result = module.analyze_detailed(module_input)

    assert result.token_audit[0].match_method is ConcretenessMatchMethod.EXACT
    assert result.token_audit[0].matched_source_term == "saw"
    assert result.token_audit[0].rating == pytest.approx(2.0)


def test_exact_two_word_expression_is_assigned_to_each_covered_token(
    tmp_path: Path,
    preprocessor,
) -> None:
    module = _module(
        tmp_path,
        [
            ("dark", 0, 1.5, 0.5, 0, 30, 1.0, 50),
            ("night", 0, 3.0, 0.5, 0, 30, 1.0, 50),
            ("dark night", 1, 4.5, 0.5, 0, 30, 1.0, 5),
        ],
    )

    result = _analyze(
        module,
        preprocessor,
        "dark night",
        configuration=ConcretenessConfiguration(exclude_proper_nouns=False),
    )
    rated = [row for row in result.token_audit if row.included]

    assert len(rated) == 2
    assert {row.match_method for row in rated} == {
        ConcretenessMatchMethod.EXACT_PHRASE
    }
    assert {row.rating for row in rated} == {4.5}
    assert len({row.match_group_id for row in rated}) == 1
    assert result.summary.matched_expression_occurrence_count == 1
    assert result.summary.statistics.mean == pytest.approx(4.5)


def test_proper_names_are_included_by_default_and_can_be_excluded(
    tmp_path: Path,
    preprocessor,
) -> None:
    module = _module(
        tmp_path,
        [
            ("alice", 0, 4.0, 0.5, 0, 30, 1.0, 50),
            ("sing", 0, 3.0, 0.5, 0, 30, 1.0, 50),
        ],
    )

    poem = preprocessor.process_document(
        create_text_document("proper-name", "Proper name", "Alice sings.")
    )
    tokens = tuple(
        replace(
            token,
            part_of_speech="PROPN",
            is_proper_noun=True,
        )
        if token.surface_form == "Alice"
        else token
        for token in poem.tokens
    )
    result = module.analyze_detailed(
        ModuleInput(
            document=poem.source,
            tokens=tokens,
            preprocessing=poem.preprocessing,
        )
    )
    alice = next(row for row in result.token_audit if row.surface_form == "Alice")

    assert alice.match_method is ConcretenessMatchMethod.EXACT
    assert alice.rating == pytest.approx(4.0)
    assert result.summary.eligible_token_count == 2

    excluded = module.analyze_detailed(
        ModuleInput(document=poem.source, tokens=tokens, preprocessing=poem.preprocessing),
        ConcretenessConfiguration(exclude_proper_nouns=True),
    )
    excluded_alice = next(row for row in excluded.token_audit if row.surface_form == "Alice")
    assert excluded_alice.match_method is ConcretenessMatchMethod.NOT_ELIGIBLE
    assert excluded_alice.rating is None
    assert "proper" in excluded_alice.reason.casefold()
    assert excluded.summary.eligible_token_count == 1


def test_empty_and_wholly_unmatched_inputs_keep_aggregates_missing(
    tmp_path: Path,
    preprocessor,
) -> None:
    module = _module(
        tmp_path,
        [("stone", 0, 4.0, 0.5, 0, 30, 1.0, 50)],
    )

    empty = _analyze(module, preprocessor, "")
    unmatched = _analyze(
        module,
        preprocessor,
        "quorvax",
        configuration=ConcretenessConfiguration(exclude_proper_nouns=False),
    )

    assert empty.summary.statistics.mean is None
    assert empty.summary.token_coverage is None
    assert empty.summary.unique_type_coverage is None
    assert unmatched.summary.statistics.mean is None
    assert unmatched.summary.token_coverage == 0.0
    assert unmatched.token_audit[0].rating is None


def test_module_contract_resource_states_and_determinism(
    tmp_path: Path,
    preprocessor,
) -> None:
    missing_spec = ResourceSpec(
        resource_id="missing-concreteness",
        display_name="Missing concreteness fixture",
        relative_path="missing.xlsx",
        version="synthetic-v1",
    )
    missing = ConcretenessModule(tmp_path, resource_spec=missing_spec)
    assert isinstance(missing, AnalysisModule)
    assert missing.validate_resources()[0].state is ResourceState.MISSING

    available = _module(
        tmp_path,
        [("café", 0, 4.2, 0.5, 0, 30, 1.0, 50)],
    )
    configuration = ConcretenessConfiguration(exclude_proper_nouns=False)
    first = _analyze(available, preprocessor, "CAFÉ", configuration=configuration)
    second = _analyze(available, preprocessor, "CAFÉ", configuration=configuration)
    assert first == second
    assert first.token_audit[0].match_method is ConcretenessMatchMethod.EXACT
    assert first.module_result.provenance.resources[0].source_sha256


def test_configuration_rejects_overlapping_or_out_of_scale_thresholds() -> None:
    with pytest.raises(ValueError, match="1-5"):
        ConcretenessConfiguration(highly_abstract_max=0.5)
    with pytest.raises(ValueError, match="below"):
        ConcretenessConfiguration(
            highly_abstract_max=4.0,
            highly_concrete_min=4.0,
        )


def test_local_supplied_concreteness_file_passes_contract_if_present() -> None:
    source = Path("resources") / BRYSBAERT_CONCRETENESS_FILENAME
    if not source.is_file():
        pytest.skip("The user-supplied concreteness workbook is not present.")

    lexicon = BrysbaertConcretenessAdapter().load(source)

    assert lexicon.validation.is_valid
    assert lexicon.validation.total_rows == 39_954
    assert lexicon.validation.usable_entries == 39_954
    assert lexicon.validation.phrase_entries == 2_896
    assert lexicon.validation.source_sha256 == BRYSBAERT_CONCRETENESS_SHA256
    assert min(entry.mean for entry in lexicon.entries.values()) == pytest.approx(1.04)
    assert max(entry.mean for entry in lexicon.entries.values()) == pytest.approx(5.0)
