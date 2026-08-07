from __future__ import annotations

import hashlib
from dataclasses import replace
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook

from versevad.adapters.kuperman_aoa import (
    REQUIRED_COLUMNS,
    KupermanAoAAdapter,
    KupermanAoAAdapterError,
)
from versevad.core import AnalysisModule, ModuleInput, ResourceSpec, ResourceState
from versevad.lexical_semantic.aoa import (
    KUPERMAN_AOA_FILENAME,
    KUPERMAN_AOA_SHA256,
    AoAConfiguration,
    AoAMatchMethod,
    AoAModule,
    attach_aoa_relationships,
)
from versevad.preprocessing import create_text_document


def _row(
    term: str,
    mean_age: float | str,
    *,
    occurrence_total: int = 20,
    numeric_response_count: int = 20,
    standard_deviation: float | str = 1.5,
    frequency_per_million: float | str = 10.0,
) -> tuple[object, ...]:
    return (
        term,
        occurrence_total,
        numeric_response_count,
        frequency_per_million,
        mean_age,
        standard_deviation,
        numeric_response_count / occurrence_total,
    )


def _write_workbook(path: Path, rows: list[tuple[object, ...]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(REQUIRED_COLUMNS)
    for row in rows:
        sheet.append(row)
    workbook.create_sheet("Sheet2")
    workbook.create_sheet("Sheet3")
    workbook.save(path)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _module(tmp_path: Path, rows: list[tuple[object, ...]]) -> AoAModule:
    source = tmp_path / "aoa.xlsx"
    _write_workbook(source, rows)
    spec = ResourceSpec(
        resource_id="synthetic-kuperman-aoa",
        display_name="Synthetic Kuperman AoA fixture",
        relative_path=source.name,
        version="synthetic-v1",
        accepted_sha256=(_sha256(source),),
        citation="Constructed test fixture.",
        license_notice="Synthetic test data.",
    )
    return AoAModule(tmp_path, resource_spec=spec)


def _analyze(
    module: AoAModule,
    preprocessor,
    text: str,
    *,
    configuration: AoAConfiguration | None = None,
):
    poem = preprocessor.process_document(
        create_text_document("aoa-test", "AoA test", text)
    )
    return module.analyze_detailed(
        ModuleInput.from_poem_document(poem),
        configuration or AoAConfiguration(),
    )


def test_adapter_preserves_source_and_retains_unavailable_values(
    tmp_path: Path,
) -> None:
    source = tmp_path / "aoa.xlsx"
    _write_workbook(
        source,
        [
            _row(
                "stone",
                4.5,
                occurrence_total=20,
                numeric_response_count=18,
                standard_deviation=1.2,
            ),
            _row(
                "wickiup",
                "NA",
                occurrence_total=21,
                numeric_response_count=0,
                standard_deviation="NA",
                frequency_per_million=0.25,
            ),
        ],
    )
    before = _sha256(source)

    lexicon = KupermanAoAAdapter().load(source)

    assert _sha256(source) == before
    assert lexicon.validation.is_valid
    assert lexicon.validation.total_rows == 2
    assert lexicon.validation.rated_entries == 1
    assert lexicon.validation.unavailable_entries == 1
    assert lexicon.validation.source_sha256 == before
    assert lexicon.lookup("stone").mean_age == pytest.approx(4.5)
    assert lexicon.lookup("stone").unknown_response_count == 2
    assert lexicon.lookup("stone").numeric_response_proportion == pytest.approx(0.9)
    assert lexicon.lookup("wickiup").mean_age is None


def test_adapter_rejects_missing_columns_duplicates_and_bad_response_ratio(
    tmp_path: Path,
) -> None:
    missing = tmp_path / "missing.xlsx"
    workbook = Workbook()
    workbook.active.title = "Sheet1"
    workbook.active.append(("Word", "Rating.Mean"))
    workbook.active.append(("stone", 4.0))
    workbook.save(missing)
    with pytest.raises(KupermanAoAAdapterError, match="expected columns"):
        KupermanAoAAdapter().load(missing)

    malformed = tmp_path / "malformed.xlsx"
    duplicate = list(_row("STONE", 30.0))
    duplicate[-1] = 0.5
    _write_workbook(malformed, [_row("stone", 4.0), tuple(duplicate)])
    with pytest.raises(KupermanAoAAdapterError) as captured:
        KupermanAoAAdapter().load(malformed)
    detail = captured.value.technical_detail.casefold()
    assert "duplicate" in detail
    assert "outside" in detail
    assert "disagrees" in detail
    assert not captured.value.data_changed


def test_exact_precedes_lemma_and_unavailable_and_unmatched_stay_missing(
    tmp_path: Path,
    preprocessor,
) -> None:
    module = _module(
        tmp_path,
        [
            _row("saw", 3.0),
            _row("see", 2.0),
            _row("stone", 4.0),
            _row(
                "wickiup",
                "NA",
                occurrence_total=20,
                numeric_response_count=0,
                standard_deviation="NA",
            ),
        ],
    )
    poem = preprocessor.process_document(
        create_text_document(
            "exact-aoa",
            "Exact AoA",
            "saw stones wickiup quorvax",
        )
    )
    tokens = tuple(
        replace(token, lemma="see", normalized_lemma="see")
        if token.surface_form == "saw"
        else token
        for token in poem.tokens
    )
    result = module.analyze_detailed(
        ModuleInput(
            document=poem.source,
            tokens=tokens,
            preprocessing=poem.preprocessing,
        ),
        AoAConfiguration(exclude_proper_nouns=False),
    )
    by_surface = {row.surface_form: row for row in result.token_audit}

    assert by_surface["saw"].match_method is AoAMatchMethod.EXACT
    assert by_surface["saw"].mean_age == pytest.approx(3.0)
    assert by_surface["stones"].match_method is AoAMatchMethod.LEMMA
    assert by_surface["stones"].mean_age == pytest.approx(4.0)
    assert by_surface["wickiup"].match_method is AoAMatchMethod.SOURCE_UNRATED
    assert by_surface["wickiup"].mean_age is None
    assert by_surface["wickiup"].matched_source_term == "wickiup"
    assert by_surface["quorvax"].match_method is AoAMatchMethod.UNMATCHED
    assert by_surface["quorvax"].mean_age is None


def test_alphabetic_number_word_matches_but_numeric_literal_is_ineligible(
    tmp_path: Path,
    preprocessor,
) -> None:
    result = _analyze(_module(tmp_path, [_row("one", 3.5)]), preprocessor, "one 27")
    by_surface = {row.surface_form: row for row in result.token_audit}

    assert by_surface["one"].included is True
    assert by_surface["one"].mean_age == pytest.approx(3.5)
    assert "alphabetically spelled" in by_surface["one"].reason
    assert by_surface["27"].eligible is False
    assert "pure numeric literal" in by_surface["27"].reason
    assert result.summary.eligible_token_count == 1
    assert result.summary.token_coverage == pytest.approx(1.0)


def test_optional_contextual_content_scope_is_not_redundant(
    tmp_path: Path,
    preprocessor,
) -> None:
    terms = (
        "the",
        "stone",
        "runs",
        "bright",
        "swiftly",
        "she",
        "can",
        "under",
        "and",
        "because",
    )
    module = _module(
        tmp_path,
        [_row(term, 3.0 + index) for index, term in enumerate(terms)],
    )
    poem = preprocessor.process_document(
        create_text_document(
            "aoa-content-scope",
            "AoA content scope",
            "the stone runs bright swiftly she can under and because",
        )
    )
    tags = {
        "the": "DET",
        "stone": "NOUN",
        "runs": "VERB",
        "bright": "ADJ",
        "swiftly": "ADV",
        "she": "PRON",
        "can": "AUX",
        "under": "ADP",
        "and": "CCONJ",
        "because": "SCONJ",
    }
    tokens = tuple(
        replace(token, part_of_speech=tags[token.normalized_form])
        for token in poem.tokens
    )
    module_input = ModuleInput(
        document=poem.source,
        tokens=tokens,
        preprocessing=poem.preprocessing,
    )

    default = module.analyze_detailed(
        module_input,
        AoAConfiguration(exclude_proper_nouns=False),
    )
    content = module.analyze_detailed(
        module_input,
        AoAConfiguration(
            exclude_proper_nouns=False,
            content_words_only=True,
        ),
    )

    assert default.summary.eligible_token_count == 10
    assert default.summary.matched_token_count == 10
    assert content.summary.eligible_token_count == 10
    assert content.summary == default.summary
    assert all(row.match_method is not AoAMatchMethod.NOT_ELIGIBLE for row in content.token_audit)
    source_notice = next(
        warning.message
        for warning in default.module_result.warnings
        if warning.code == "source_sampling_and_context"
    )
    assert "polyfunctional" in source_notice


def test_proper_names_are_included_by_default_and_can_be_excluded(
    tmp_path: Path,
    preprocessor,
) -> None:
    module = _module(tmp_path, [_row("alice", 5.0), _row("sings", 4.0)])
    poem = preprocessor.process_document(
        create_text_document("proper-aoa", "Proper AoA", "Alice sings")
    )
    tokens = tuple(
        replace(token, part_of_speech="PROPN", is_proper_noun=True)
        if token.surface_form == "Alice"
        else token
        for token in poem.tokens
    )
    result = module.analyze_detailed(
        ModuleInput(
            document=poem.source,
            tokens=tokens,
            preprocessing=poem.preprocessing,
        )
    )
    alice = next(row for row in result.token_audit if row.surface_form == "Alice")

    assert alice.match_method is AoAMatchMethod.EXACT
    assert alice.mean_age == pytest.approx(5.0)
    assert result.summary.eligible_token_count == 2

    excluded = module.analyze_detailed(
        ModuleInput(
            document=poem.source,
            tokens=tokens,
            preprocessing=poem.preprocessing,
        ),
        AoAConfiguration(exclude_proper_nouns=True),
    )
    excluded_alice = next(
        row for row in excluded.token_audit if row.surface_form == "Alice"
    )
    assert excluded_alice.match_method is AoAMatchMethod.NOT_ELIGIBLE
    assert excluded_alice.mean_age is None
    assert "proper" in excluded_alice.reason.casefold()
    assert excluded.summary.eligible_token_count == 1


def test_repetition_statistics_bands_structure_and_response_evidence(
    tmp_path: Path,
    preprocessor,
) -> None:
    module = _module(
        tmp_path,
        [
            _row("early", 3.0, numeric_response_count=4),
            _row("middle", 8.0),
            _row("later", 14.0),
        ],
    )
    result = _analyze(
        module,
        preprocessor,
        "early early\nmiddle\n\nlater",
        configuration=AoAConfiguration(exclude_proper_nouns=False),
    )

    assert result.summary.statistics.mean == pytest.approx(7.0)
    assert result.summary.statistics.median == pytest.approx(5.5)
    assert result.summary.interquartile_range == pytest.approx(6.5)
    assert result.summary.low_response_token_count == 2
    assert result.summary.minimum_source_numeric_responses == 4
    bands = {band.band_id: band for band in result.summary.bands}
    assert bands["early_acquired"].token_count == 2
    assert bands["middle_range"].token_count == 1
    assert bands["later_acquired"].token_count == 1
    assert len(result.line_summaries) == 4
    assert result.line_summaries[2].statistics.mean is None
    assert len(result.stanza_summaries) == 2
    early = next(term for term in result.term_summaries if term.source_term == "early")
    assert early.matched_token_occurrences == 2
    assert result.earliest_acquired_terms[0].source_term == "early"
    assert result.latest_acquired_terms[0].source_term == "later"


def test_empty_unmatched_unicode_resource_states_and_determinism(
    tmp_path: Path,
    preprocessor,
) -> None:
    missing_spec = ResourceSpec(
        resource_id="missing-aoa",
        display_name="Missing AoA fixture",
        relative_path="missing.xlsx",
        version="synthetic-v1",
    )
    missing = AoAModule(tmp_path, resource_spec=missing_spec)
    assert isinstance(missing, AnalysisModule)
    assert missing.validate_resources()[0].state is ResourceState.MISSING

    module = _module(tmp_path, [_row("caf\u00e9", 4.2)])
    configuration = AoAConfiguration(exclude_proper_nouns=False)
    first = _analyze(module, preprocessor, "CAF\u00c9", configuration=configuration)
    second = _analyze(module, preprocessor, "CAF\u00c9", configuration=configuration)
    empty = _analyze(module, preprocessor, "")
    unmatched = _analyze(
        module,
        preprocessor,
        "quorvax",
        configuration=configuration,
    )

    assert first == second
    assert first.token_audit[0].match_method is AoAMatchMethod.EXACT
    assert empty.summary.statistics.mean is None
    assert empty.summary.token_coverage is None
    assert unmatched.summary.statistics.mean is None
    assert unmatched.summary.token_coverage == 0.0
    assert unmatched.token_audit[0].mean_age is None


def test_relationships_use_unique_surface_types_and_require_three_pairs(
    tmp_path: Path,
    preprocessor,
) -> None:
    module = _module(
        tmp_path,
        [_row("alpha", 2.0), _row("beta", 4.0), _row("gamma", 6.0)],
    )
    aoa = _analyze(
        module,
        preprocessor,
        "alpha alpha beta gamma",
        configuration=AoAConfiguration(exclude_proper_nouns=False),
    )
    values_by_surface = {
        "alpha": (6.0, 1.0),
        "beta": (4.0, 2.0),
        "gamma": (2.0, 3.0),
    }
    frequency_rows = []
    concreteness_rows = []
    for row in aoa.token_audit:
        zipf, concrete = values_by_surface[row.normalized_form]
        frequency_rows.append(
            SimpleNamespace(
                token_id=row.token_id,
                included=True,
                zipf_value=zipf,
            )
        )
        concreteness_rows.append(
            SimpleNamespace(
                token_id=row.token_id,
                included=True,
                rating=concrete,
                source_is_multiword=False,
            )
        )
    frequency = SimpleNamespace(
        module_result=SimpleNamespace(module_name="lexical_frequency"),
        token_audit=tuple(frequency_rows),
    )
    concreteness = SimpleNamespace(
        module_result=SimpleNamespace(module_name="concreteness"),
        token_audit=tuple(concreteness_rows),
    )

    related = attach_aoa_relationships(
        aoa,
        frequency=frequency,
        concreteness=concreteness,
    )
    by_id = {item.relationship_id: item for item in related.relationships}

    assert by_id["aoa_vs_frequency"].pair_count == 3
    assert by_id["aoa_vs_frequency"].coefficient == pytest.approx(-1.0)
    assert by_id["aoa_vs_concreteness"].coefficient == pytest.approx(1.0)
    assert len(related.module_result.metrics) == len(aoa.module_result.metrics) + 2


def test_configuration_rejects_overlapping_bands_and_sparse_relationships() -> None:
    with pytest.raises(ValueError, match="below"):
        AoAConfiguration(early_acquired_max=12.0, later_acquired_min=12.0)
    with pytest.raises(ValueError, match="three"):
        AoAConfiguration(minimum_relationship_types=2)


def test_local_official_kuperman_file_passes_contract_if_present() -> None:
    source = Path("resources") / KUPERMAN_AOA_FILENAME
    if not source.is_file():
        pytest.skip("The official local Kuperman workbook is not present.")

    lexicon = KupermanAoAAdapter().load(source)

    assert lexicon.validation.is_valid
    assert lexicon.validation.total_rows == 31_124
    assert lexicon.validation.rated_entries == 31_105
    assert lexicon.validation.unavailable_entries == 19
    assert lexicon.validation.source_sha256 == KUPERMAN_AOA_SHA256
    assert lexicon.lookup("the").mean_age == pytest.approx(3.983747)
    assert lexicon.lookup("and").mean_age == pytest.approx(4.569882)
    assert lexicon.lookup("he").mean_age == pytest.approx(3.813235)
    assert lexicon.lookup("of").mean_age == pytest.approx(4.548568)
    assert lexicon.lookup("to").mean_age == pytest.approx(3.951776)
    assert lexicon.lookup("wickiup").mean_age is None
